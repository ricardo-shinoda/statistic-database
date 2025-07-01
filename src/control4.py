from googleapiclient.http import MediaFileUpload
from google.oauth2 import service_account
from googleapiclient.discovery import build
import openpyxl
import pandas as pd
import json
import zipfile
import shutil
import os
import glob
from decouple import config
from dotenv import load_dotenv

# Path to the downloaded zip file
zip_file_path = "/home/ricardo/Downloads/Fatura-CPF.zip"

# Password for the zip file
zip_password = config('ZIP_PASSWORD')

# Extract the contents of the zip file
extracted_folder = "/home/ricardo/Downloads/invoice"
with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
    zip_ref.extractall(extracted_folder, pwd=bytes(zip_password, 'utf-8'))

# Find the path of the CSV file inside the extracted folder
csv_files = glob.glob(os.path.join(extracted_folder, '*.csv'))

if csv_files:
    # Use the first CSV file found (you may want to implement specific logic if there are multiple CSV files)
    extracted_text_file = csv_files[0]

    # Path to the Text Editor app in Pop OS
    text_editor_path = "/usr/bin/gedit"

source_directory = "/home/ricardo/Downloads/invoice"

# Rename it as needed
file = "invoice"

target_extension = ".csv"

destination_directory = "/home/ricardo/programing/statistic/src"

destination_file_path = None  # Initialize the variable outside the loop

for root, dirs, files in os.walk(source_directory):
    for current_file in files:  # Use a different variable name for the loop
        if current_file.endswith(target_extension):
            source_file_path = os.path.join(root, current_file)
            destination_file_name = 'invoice.csv'
            destination_file_path = os.path.join(
                destination_directory, destination_file_name)
            shutil.move(source_file_path, destination_file_path)
            break

if destination_file_path:
    df = pd.read_csv(destination_file_path, delimiter=';')

    data = df.to_dict(orient='records')


#! rename this variable to save the file according to the invoice month
month = "2025-06"

# Read the CSV file with the specified delimiter
df = pd.read_csv('invoice.csv', delimiter=';')

# Print column names and first few rows for debugging
print("Column Names:", df.columns)
print("First Few Rows:\n", df.head())

# Exclude rows with "Inclusao de Pagamento" in the "Descrição" column
df = df[df['Descrição'] != 'Inclusao de Pagamento    ']

# Load the description mapping from the JSON file
with open('/home/ricardo/programing/statistic/src/description.json') as f:
    description_list = json.load(f)

# Convert the list of dictionaries to a dictionary
description_mapping = {item['original']: item['new']
                       for item in description_list}

# Create a new column "Mapped_Categoria" based on the mapping
df['Mapped_Categoria'] = df['Descrição'].map(description_mapping)

# Use the original "Categoria" if there is no mapping, otherwise use the mapped value
df['Categoria'] = df['Mapped_Categoria'].combine_first(df['Categoria'])

# Drop the temporary "Mapped_Categoria" column
df = df.drop(columns=['Mapped_Categoria'])

# Remove leading and trailing whitespaces from column names
df.columns = df.columns.str.strip()

# Print column names and first few rows for debugging
print("Column Names after processing:\n", df.columns)
print("First Few Rows after processing:\n", df.head())

# Group by "Categoria" and calculate the sum of "Valor (em R$)"
try:
    category_sum = df.groupby('Categoria')['Valor (em R$)'].sum().reset_index()
except KeyError as e:
    print(
        f"KeyError: {e}. Check if the column 'Valor (em R$)' is present in the DataFrame.")
    # Add additional debugging information if needed

# Group by "Nome no Cartão" and calculate the sum of "Valor (em R$)"
try:
    user_sum = df.groupby('Nome no Cartão')[
        'Valor (em R$)'].sum().reset_index()
    # Calculate the total amount of expenses
    total_expenses = user_sum['Valor (em R$)'].sum()
except KeyError as e:
    print(
        f"KeyError: {e}. Check if the column 'Valor (em R$)' is present in the DataFrame.")

# Save both DataFrames to a single sheet in a new Excel file, with the summary table on the right of the original data
with pd.ExcelWriter(f'/home/ricardo/programing/statistic/src/credit_card/xlsx/{month}.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Data', index=False, startrow=0, startcol=0)
    category_sum.to_excel(writer, sheet_name='Data', index=False,
                          startrow=0, startcol=df.shape[1] + 1, header=True)
    user_sum.to_excel(writer, sheet_name='Data', index=False,
                      startrow=0, startcol=df.shape[1] + 3, header=True)

    # Create a DataFrame for total expenses
    total_expenses_df = pd.DataFrame(
        {'Nome no Cartão': ['Total Expenses'], 'Valor (em R$)': [total_expenses]})
    # Append the total expenses to the Excel file
    total_expenses_df.to_excel(writer, sheet_name='Data', index=False,
                               startrow=user_sum.shape[0] + 2, startcol=df.shape[1] + 3, header=True)

df = pd.read_csv('invoice.csv', delimiter=';')

data = df.to_dict(orient='records')

with open(f'/home/ricardo/programing/statistic/src/credit_card/json/{month}.json', 'w', encoding='utf-8') as json_file:
    json.dump(data, json_file, indent=4, ensure_ascii=False)

print("Conversion complete. Data saved on 'output.json'.")

print(f'Conversion complete, Data and Summary saved in {month}.xlsx.')


# Save it as new tab on Controle


file_name = month

# Especificação dos arquivos de origem e destino
source_file_path = f'/home/ricardo/programing/statistic/src/credit_card/xlsx/{file_name}.xlsx'
target_file_path = '/home/ricardo/programing/statistic/src/Controle.xlsx'

# Nome da aba de origem e nome da nova aba de destino
source_sheet_name = 'Data'
new_target_sheet_name = f'{file_name}'  # Change the sheet name to 'Controle2'

# Carrega a aba de origem usando openpyxl
source_workbook = openpyxl.load_workbook(source_file_path, data_only=True)
source_sheet = source_workbook[source_sheet_name]

# Carrega a planilha de destino (ou cria uma nova se não existir)
if os.path.exists(target_file_path):
    target_workbook = openpyxl.load_workbook(target_file_path)
else:
    target_workbook = openpyxl.Workbook()

# Create a new sheet named 'Controle2' or get it if it already exists
if new_target_sheet_name in target_workbook.sheetnames:
    target_sheet = target_workbook[new_target_sheet_name]
else:
    target_sheet = target_workbook.create_sheet(title=new_target_sheet_name)

# Copia os dados da aba de origem para a aba de destino
for row in source_sheet.iter_rows(values_only=True):
    target_sheet.append(row)

# Salva a planilha de destino
target_workbook.save(target_file_path)

# Faz o upload da planilha para o Google Drive

SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = 'service.account.json'
PARENT_FOLDER_ID = config('PARENT_FOLDER_ID_DB')


def authenticate():
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return creds


def upload_file(file_path):
    creds = authenticate()
    service = build('drive', 'v3', credentials=creds)

    file_metadata = {
        'name': f'Controle_{month}.xlsx',
        # 'name': f'Controle.xlsx',
        'parents': [PARENT_FOLDER_ID]
    }

    media_body = MediaFileUpload(file_path, resumable=True)

    file = service.files().create(
        body=file_metadata,
        media_body=media_body
    ).execute()


# Upload the target file to Google Drive
upload_file(target_file_path)
