# This scrip take the .xlsx file saved on credir_card folder and move it to Controle and move it to G-Drive

import openpyxl
import os
from googleapiclient.discovery import build
from google.oauth2 import service_account
from googleapiclient.http import MediaFileUpload

file_name = '2022-12'

# Especificação dos arquivos de origem e destino
source_file_path = f'/home/ricardo/code/statistic/src/credit_card/xlsx/{
    file_name}.xlsx'
target_file_path = '/home/ricardo/code/statistic/src/Controle.xlsx'

# Nome da aba de origem e nome da nova aba de destino
source_sheet_name = 'Sheet1'
new_target_sheet_name = f'{file_name}'  # Change the sheet name to 'Controle2'

# Carrega a aba de origem usando openpyxl
source_workbook = openpyxl.load_workbook(source_file_path, data_only=True)
source_sheet = source_workbook[source_sheet_name]

# Carrega a planilha de destino (ou cria uma nova se não existir)
if os.path.exists(target_file_path):
    target_workbook = openpyxl.load_workbook(target_file_path)
else:
    target_workbook = openpyxl.Workbook()

# Create a new sheet named 'Controle2' or get it if it already exists
if new_target_sheet_name in target_workbook.sheetnames:
    target_sheet = target_workbook[new_target_sheet_name]
else:
    target_sheet = target_workbook.create_sheet(title=new_target_sheet_name)

# Copia os dados da aba de origem para a aba de destino
for row in source_sheet.iter_rows(values_only=True):
    target_sheet.append(row)

# Salva a planilha de destino
target_workbook.save(target_file_path)

# Faz o upload da planilha para o Google Drive

SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = 'service.account.json'
PARENT_FOLDER_ID = "1qyMZb6P7H5oaq2zxoaqs17VJWSnHNgnG"


def authenticate():
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return creds


def upload_file(file_path):
    creds = authenticate()
    service = build('drive', 'v3', credentials=creds)

    file_metadata = {
        'name': 'Controle.xlsx',
        'parents': [PARENT_FOLDER_ID]
    }

    media_body = MediaFileUpload(file_path, resumable=True)

    file = service.files().create(
        body=file_metadata,
        media_body=media_body
    ).execute()


# Upload the target file to Google Drive
upload_file(target_file_path)
