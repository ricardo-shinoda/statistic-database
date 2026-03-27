# THIS CODE IS WORKING WITH AN ISSUE ON THE DATA, NEED TO REMOVE \

import pandas as pd
import json
import openpyxl
import os
from googleapiclient.http import MediaFileUpload
from google.oauth2 import service_account
from googleapiclient.discovery import build

# Define file paths
csv_file_path = "/home/ricardo/Downloads/invoice/invoice.csv"  # Path to the CSV file
month = "202301TESTE"  # You can update this dynamically as needed

json_file_path = f"/home/ricardo/code/statistic/src/credit_card/json/{month}.json"
xlsx_file_path = f"/home/ricardo/code/statistic/src/credit_card/xlsx/{month}.xlsx"
target_file_path = '/home/ricardo/code/statistic/src/Controle.xlsx'

# Read the CSV file with the specified delimiter
df = pd.read_csv(csv_file_path, delimiter=';')

# Exclude rows with "Inclusao de Pagamento" in the "Descrição" column
df = df[df['Descrição'] != 'Inclusao de Pagamento    ']

# Load the description mapping from the JSON file
with open('/home/ricardo/code/statistic/src/description.json') as f:
    description_list = json.load(f)

# Convert the list of dictionaries to a dictionary for mapping
description_mapping = {item['original']: item['new']
                       for item in description_list}

# Apply the mapping and fill in categories
df['Mapped_Categoria'] = df['Descrição'].map(description_mapping)
df['Categoria'] = df['Mapped_Categoria'].combine_first(df['Categoria'])
df = df.drop(columns=['Mapped_Categoria'])
df.columns = df.columns.str.strip()

# Group by "Categoria" and sum "Valor (em R$)"
category_sum = df.groupby('Categoria')['Valor (em R$)'].sum().reset_index()

# Group by "Nome no Cartão" and sum "Valor (em R$)"
user_sum = df.groupby('Nome no Cartão')['Valor (em R$)'].sum().reset_index()

# Calculate total expenses
total_expenses = user_sum['Valor (em R$)'].sum()

# Save the data to Excel
with pd.ExcelWriter(xlsx_file_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Data', index=False, startrow=0, startcol=0)
    category_sum.to_excel(writer, sheet_name='Data', index=False,
                          startrow=0, startcol=df.shape[1] + 1, header=True)
    user_sum.to_excel(writer, sheet_name='Data', index=False,
                      startrow=0, startcol=df.shape[1] + 3, header=True)
    total_expenses_df = pd.DataFrame(
        {'Nome no Cartão': ['Total Expenses'], 'Valor (em R$)': [total_expenses]})
    total_expenses_df.to_excel(writer, sheet_name='Data', index=False,
                               startrow=user_sum.shape[0] + 2, startcol=df.shape[1] + 3, header=True)

# Save the data to JSON
df.to_json(json_file_path, orient='records', indent=4, force_ascii=False)

# Load the target Excel file and add the new data as a sheet
source_sheet_name = 'Data'
new_target_sheet_name = month  # New sheet name based on the month

if os.path.exists(target_file_path):
    target_workbook = openpyxl.load_workbook(target_file_path)
else:
    target_workbook = openpyxl.Workbook()

if new_target_sheet_name in target_workbook.sheetnames:
    target_sheet = target_workbook[new_target_sheet_name]
else:
    target_sheet = target_workbook.create_sheet(title=new_target_sheet_name)

source_workbook = openpyxl.load_workbook(xlsx_file_path, data_only=True)
source_sheet = source_workbook[source_sheet_name]

# Copy data from source to target sheet
for row in source_sheet.iter_rows(values_only=True):
    target_sheet.append(row)

target_workbook.save(target_file_path)

# Google Drive upload
SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = 'service.account.json'
PARENT_FOLDER_ID = ('PARENT_FOLDER_ID_DB')


def authenticate():
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return creds


def upload_file(file_path):
    creds = authenticate()
    service = build('drive', 'v3', credentials=creds)
    file_metadata = {'name': f'Controle_{month}.xlsx',
                     'parents': [PARENT_FOLDER_ID]}
    media_body = MediaFileUpload(file_path, resumable=True)
    service.files().create(body=file_metadata, media_body=media_body).execute()


# Upload the file to Google Drive
upload_file(target_file_path)

print(f"CSV converted and saved as JSON and Excel. Data uploaded to Google Drive.")
